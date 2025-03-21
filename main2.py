import cv2
import tkinter as tk
from tkinter import messagebox, simpledialog
from simple_facerec import SimpleFacerec
from datetime import datetime
import openpyxl
import os
from PIL import Image, ImageTk

# Initialize attendance sheet
attendance_file = "attendance.xlsx"

def initialize_attendance_file():
    try:
        # Load existing workbook
        workbook = openpyxl.load_workbook(attendance_file)
        sheet = workbook.active
    except FileNotFoundError:
        # Create new workbook if file doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Date", "Time"])
        workbook.save(attendance_file)

# Mark attendance in the Excel sheet
def mark_attendance(name):
    # Get the current date and time
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    # Load the workbook
    workbook = openpyxl.load_workbook(attendance_file)
    sheet = workbook.active

    # Check if the name is already marked for today
    for row in sheet.iter_rows(values_only=True):
        if row[0] == name and row[1] == current_date:
            return  # Already marked attendance

    # Append the new attendance record
    sheet.append([name, current_date, current_time])
    workbook.save(attendance_file)

# Encode faces from a folder
sfr = SimpleFacerec()
sfr.load_encoding_images("images/")

# Create main application window
root = tk.Tk()
root.title("Face Recognition Attendance System")

# Set window size
root.geometry("1000x600")

# Style Configuration
root.config(bg="#f0f0f0")

# Initialize the label to display the video feed
video_label = tk.Label(root, bg="white")
video_label.pack(padx=10, pady=10)

# Label for status updates (e.g., attendance marked)
status_label = tk.Label(root, text="Status: Waiting for face...", font=("Arial", 14), bg="#f0f0f0")
status_label.pack(pady=10)

# Initialize camera
cap = cv2.VideoCapture(0)

# Initialize attendance tracker for the session
marked_names = set()

# Function to update the UI with new frames
def update_frame():
    ret, frame = cap.read()
    if ret:
        # Detect Faces
        face_locations, face_names = sfr.detect_known_faces(frame)
        
        for face_loc, name in zip(face_locations, face_names):
            y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]
            
            # Display name and bounding box on the frame
            cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 200), 2)
            cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 200), 4)
            
            # Mark attendance
            if name != "Unknown" and name not in marked_names:
                mark_attendance(name)
                marked_names.add(name)
                status_label.config(text=f"Status: {name} marked attendance")

        # Convert image to PhotoImage for displaying on Tkinter
        frame_rgb = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(frame_rgb)
        img_tk = ImageTk.PhotoImage(img)

        # Update the label with the new image
        video_label.img_tk = img_tk
        video_label.config(image=img_tk)
    
    # Continue the frame update loop
    video_label.after(10, update_frame)

# Function to start the camera stream
def start_camera():
    status_label.config(text="Status: Camera started")
    update_frame()

# Function to stop the camera stream
def stop_camera():
    status_label.config(text="Status: Camera stopped")
    cap.release()
    cv2.destroyAllWindows()

# Function to register a new face (single capture)
def register_face():
    # Ask user for their name
    name = simpledialog.askstring("Input", "Enter your name:")
    if not name:
        messagebox.showwarning("Input Error", "Name is required to register.")
        return

    # Ensure the name is valid for file names (e.g., no special characters)
    name = name.replace(" ", "_").replace("/", "_").replace("\\", "_")

    # Capture face image and save it directly into the 'images' folder
    ret, frame = cap.read()
    if ret:
        # Convert the frame to grayscale
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
        
        # Detect faces in the frame
        face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")
        faces = face_cascade.detectMultiScale(gray, 1.3, 5)
        
        if len(faces) > 0:
            # Save the captured face image to the 'images' folder
            (x, y, w, h) = faces[0]  # Take the first detected face
            face_img = frame[y:y+h, x:x+w]
            face_filename = f"images/{name}.jpg"
            cv2.imwrite(face_filename, face_img)

            # Update status
            messagebox.showinfo("Registration Complete", f"Face image captured for {name}.")
        else:
            messagebox.showwarning("No Face Detected", "No face detected. Please try again.")
    else:
        messagebox.showerror("Error", "Failed to capture image.")

    # Reload the face encodings after new image is added
    sfr.load_encoding_images("images/")

# Start/Stop Camera Button
start_button = tk.Button(root, text="Start Camera", font=("Arial", 12), bg="#4CAF50", fg="white", command=start_camera)
start_button.pack(pady=5, side="left", padx=10)

stop_button = tk.Button(root, text="Stop Camera", font=("Arial", 12), bg="#f44336", fg="white", command=stop_camera)
stop_button.pack(pady=5, side="right", padx=10)

# Register New Face Button
register_button = tk.Button(root, text="Register New Face", font=("Arial", 12), bg="#008CBA", fg="white", command=register_face)
register_button.pack(pady=5)

# Initialize attendance file and UI
initialize_attendance_file()

# Start the Tkinter event loop
root.mainloop()
