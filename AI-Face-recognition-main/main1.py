import cv2
from simple_facerec import SimpleFacerec
import openpyxl
from datetime import datetime

# Initialize attendance sheet
attendance_file = "attendance.xlsx"
def initialize_attendance_file():
    try:
        # Load existing workbook
        workbook = openpyxl.load_workbook(attendance_file)
        sheet = workbook.active
    except FileNotFoundError:
        # Create new workbook if file doesn't exist
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Attendance"
        sheet.append(["Name", "Date", "Time"])
        workbook.save(attendance_file)

# Mark attendance in the Excel sheet
def mark_attendance(name):
    # Get the current date and time
    now = datetime.now()
    current_date = now.strftime("%Y-%m-%d")
    current_time = now.strftime("%H:%M:%S")

    # Load the workbook
    workbook = openpyxl.load_workbook(attendance_file)
    sheet = workbook.active

    # Check if the name is already marked for today
    for row in sheet.iter_rows(values_only=True):
        if row[0] == name and row[1] == current_date:
            return  # Already marked attendance

    # Append the new attendance record
    sheet.append([name, current_date, current_time])
    workbook.save(attendance_file)

# Encode faces from a folder
sfr = SimpleFacerec()
sfr.load_encoding_images("images/")

# Load Camera
cap = cv2.VideoCapture(0)

initialize_attendance_file()

while True:
    ret, frame = cap.read()

    # Detect Faces
    face_locations, face_names = sfr.detect_known_faces(frame)
    for face_loc, name in zip(face_locations, face_names):
        y1, x2, y2, x1 = face_loc[0], face_loc[1], face_loc[2], face_loc[3]

        # Display name and bounding box on the frame
        cv2.putText(frame, name, (x1, y1 - 10), cv2.FONT_HERSHEY_DUPLEX, 1, (0, 0, 200), 2)
        cv2.rectangle(frame, (x1, y1), (x2, y2), (0, 0, 200), 4)

        # Mark attendance
        if name != "Unknown":
            mark_attendance(name)

    # Display the frame in a resized window for easier viewing
    resized_frame = cv2.resize(frame, (800, 600))
    cv2.imshow("Frame", resized_frame)

    key = cv2.waitKey(1)
    if key == 27:  # Escape key to stop
        break

cap.release()
cv2.destroyAllWindows()